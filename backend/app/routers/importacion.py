import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import calculations, models
from ..database import get_db

router = APIRouter(prefix="/importacion", tags=["Importación"])

# La búsqueda de productos existentes se hace por Nombre (case-insensitive, sin espacios
# al principio/final). No requiere "código de producto": con ~100 SKU y una sola persona
# armando la planilla, es más práctico. El campo `codigo` existe en el modelo para uso futuro
# (código de barras, disambiguación) pero hoy no se usa en la importación.
COLUMNAS = ["Producto", "Categoria", "Costo", "Cantidad", "Descuento", "FechaCompra", "PrecioVenta"]
COLUMNAS_OBLIGATORIAS = ["Producto", "Costo", "Cantidad"]


def _norm(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _parse_fecha(valor):
    if valor in (None, ""):
        return date.today()
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(valor).strip(), fmt).date()
        except ValueError:
            continue
    return date.today()


@router.get("/plantilla")
def descargar_plantilla(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importación"

    # Una columna por cada Atributo que exista HOY en el sistema (Talle, Color, etc.) — se arma
    # en el momento de la descarga, así que si se agrega un atributo nuevo desde la pantalla
    # Atributos, la próxima plantilla ya lo incluye sin tocar código.
    atributos = db.query(models.Atributo).order_by(models.Atributo.id).all()
    ws.append(COLUMNAS + [a.nombre for a in atributos])

    def valor_ejemplo(atributo, indice):
        valores = atributo.valores
        return valores[indice % len(valores)].valor if valores else ""

    ws.append(["Top Básico", "Remeras", 15000, 10, "", "2026-07-01", 32000] + ["" for _ in atributos])
    ws.append(["Vestido Fiesta", "Vestidos", 34000, 5, 10, "2026-07-01", ""] + ["" for _ in atributos])
    ws.append(["Top Básico", "Remeras", 16000, 5, "", "2026-07-15", ""] + ["" for _ in atributos])

    if atributos:
        # 1-2 filas de ejemplo con atributos completos, usando valores REALES ya cargados en
        # valores_atributo (no inventados) — muestran cómo reponer/dar de alta con variantes.
        ws.append(
            ["Calza Deportiva", "Calzas", 12000, 3, "", "2026-07-01", 25000]
            + [valor_ejemplo(a, 0) for a in atributos]
        )
        ws.append(
            ["Calza Deportiva", "Calzas", 12000, 2, "", "2026-07-01", ""]
            + [valor_ejemplo(a, 1) for a in atributos]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_importacion_fashbalance.xlsx"},
    )


@router.post("/procesar")
async def procesar(file: UploadFile = File(...), db: Session = Depends(get_db)):
    umbral_cambio_costo_pct = float(calculations.get_configuracion(db).umbral_cambio_costo_pct)

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "El archivo debe ser una planilla Excel (.xlsx).")

    contenido = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo. Verificá que sea un .xlsx válido.")

    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(400, "La planilla está vacía.")

    encabezado = [_norm(c) for c in filas[0]]
    idx = {col: (encabezado.index(col) if col in encabezado else None) for col in COLUMNAS}

    faltantes = [c for c in COLUMNAS_OBLIGATORIAS if idx[c] is None]
    if faltantes:
        raise HTTPException(400, f"Faltan columnas obligatorias en la planilla: {', '.join(faltantes)}.")

    # Cualquier columna que no sea una de las fijas se interpreta como un atributo (Talle, Color,
    # etc.). Solo se reconocen atributos que YA existen en el sistema — si hay alguna columna que no
    # matchea ningún atributo existente, se cancela TODA la importación antes de escribir nada (esto
    # es un error estructural de la planilla, no de una fila puntual).
    columnas_fijas_norm = {c.lower() for c in COLUMNAS}
    atributos_por_nombre = {a.nombre.strip().lower(): a for a in db.query(models.Atributo).all()}
    columnas_atributo = []  # [(indice_columna, Atributo), ...] en el orden en que aparecen en el encabezado
    columnas_desconocidas = []
    for i, nombre_col in enumerate(encabezado):
        if not nombre_col or nombre_col.lower() in columnas_fijas_norm:
            continue
        atributo = atributos_por_nombre.get(nombre_col.lower())
        if atributo is None:
            columnas_desconocidas.append(nombre_col)
        else:
            columnas_atributo.append((i, atributo))

    if columnas_desconocidas:
        disponibles = ", ".join(a.nombre for a in atributos_por_nombre.values()) or "ninguno todavía"
        raise HTTPException(
            400,
            "La planilla tiene columnas que no corresponden a ningún atributo existente: "
            f"{', '.join(columnas_desconocidas)}. Atributos disponibles en el sistema: {disponibles}. "
            "Si es un atributo nuevo, crealo primero desde la pantalla Atributos.",
        )

    categorias_cache = {c.nombre.strip().lower(): c for c in db.query(models.Categoria).all()}
    productos_cache = {p.nombre.strip().lower(): p for p in db.query(models.Producto).all()}
    valores_por_atributo_cache: dict[int, dict] = {}

    productos_creados = []
    compras_registradas = []
    cambios_costo = []
    errores = []
    productos_nuevos_ids = []

    def val(fila, col):
        i = idx.get(col)
        return fila[i] if i is not None and i < len(fila) else None

    def valor_columna(fila, col_index):
        return fila[col_index] if col_index is not None and col_index < len(fila) else None

    def valor_atributo_por_texto(atributo_id, texto_norm):
        if atributo_id not in valores_por_atributo_cache:
            valores_por_atributo_cache[atributo_id] = {
                v.valor.strip().lower(): v
                for v in db.query(models.ValorAtributo).filter(models.ValorAtributo.atributo_id == atributo_id).all()
            }
        return valores_por_atributo_cache[atributo_id].get(texto_norm)

    def descripcion_variante(variante_id):
        if variante_id is None:
            return None
        filas_desc = (
            db.query(models.ValorAtributo.valor, models.ProductoAtributo.orden)
            .join(models.VarianteValor, models.VarianteValor.valor_atributo_id == models.ValorAtributo.id)
            .join(models.Variante, models.VarianteValor.variante_id == models.Variante.id)
            .join(
                models.ProductoAtributo,
                (models.ProductoAtributo.producto_id == models.Variante.producto_id)
                & (models.ProductoAtributo.atributo_id == models.ValorAtributo.atributo_id),
            )
            .filter(models.VarianteValor.variante_id == variante_id)
            .order_by(models.ProductoAtributo.orden)
            .all()
        )
        return " / ".join(v for v, _ in filas_desc)

    def resolver_variante_para_fila(producto, atributos_fila):
        """Devuelve (variante_id, error). Si el producto no tiene (ni necesita) variantes para esta
        fila, devuelve (None, None). Si es la primera vez que este producto ve atributos (recién
        creado, o recién activado sobre la marcha), configura sus `producto_atributos` a partir de
        lo que trae ESTA fila (orden = posición de columna en la planilla). Si el producto YA tenía
        atributos configurados (por Catálogo o por una fila anterior de este mismo import), exige que
        la fila traiga valor para todos ellos; columnas de atributo ajenas a su configuración se
        ignoran."""
        configurados = (
            db.query(models.ProductoAtributo)
            .filter(models.ProductoAtributo.producto_id == producto.id)
            .order_by(models.ProductoAtributo.orden)
            .all()
        )
        if not configurados:
            if not atributos_fila:
                return None, None
            producto.tiene_variantes = True
            atributo_ids_usados = set(atributos_fila.keys())
            orden = 1
            for _, atributo in columnas_atributo:
                if atributo.id in atributo_ids_usados:
                    db.add(models.ProductoAtributo(producto_id=producto.id, atributo_id=atributo.id, orden=orden))
                    orden += 1
            db.flush()
            configurados = (
                db.query(models.ProductoAtributo)
                .filter(models.ProductoAtributo.producto_id == producto.id)
                .order_by(models.ProductoAtributo.orden)
                .all()
            )

        valor_ids = []
        for pa in configurados:
            va = atributos_fila.get(pa.atributo_id)
            if va is None:
                atributo = db.get(models.Atributo, pa.atributo_id)
                return None, f"Este producto tiene variantes: falta indicar valor de '{atributo.nombre}' para esta fila."
            valor_ids.append(va.id)

        variante = calculations.resolver_o_crear_variante(db, producto.id, valor_ids)
        return variante.id, None

    for n, fila in enumerate(filas[1:], start=2):
        nombre = _norm(val(fila, "Producto"))
        if not nombre:
            continue  # fila vacía, se ignora sin marcar error

        try:
            costo_raw = val(fila, "Costo")
            cantidad_raw = val(fila, "Cantidad")
            if costo_raw in (None, "") or cantidad_raw in (None, ""):
                errores.append({"fila": n, "producto": nombre, "motivo": "Falta Costo o Cantidad."})
                continue

            costo = Decimal(str(costo_raw))
            cantidad = int(cantidad_raw)
            if cantidad <= 0:
                errores.append({"fila": n, "producto": nombre, "motivo": "La cantidad debe ser mayor a 0."})
                continue

            atributos_fila = {}
            error_atributo = None
            for col_index, atributo in columnas_atributo:
                crudo = _norm(valor_columna(fila, col_index))
                if not crudo:
                    continue
                va = valor_atributo_por_texto(atributo.id, crudo.lower())
                if va is None:
                    error_atributo = (
                        f"El valor '{crudo}' no existe para el atributo '{atributo.nombre}' "
                        "(revisá mayúsculas/tildes o cargalo primero en Atributos)."
                    )
                    break
                atributos_fila[atributo.id] = va
            if error_atributo:
                errores.append({"fila": n, "producto": nombre, "motivo": error_atributo})
                continue

            descuento_raw = val(fila, "Descuento")
            if descuento_raw not in (None, ""):
                descuento_pct = Decimal(str(descuento_raw))
                costo_final = costo * (Decimal("1") - descuento_pct / Decimal("100"))
            else:
                costo_final = costo
            costo_final = round(costo_final, 2)

            fecha = _parse_fecha(val(fila, "FechaCompra"))
            clave = nombre.lower()
            producto = productos_cache.get(clave)

            if producto is None:
                # Producto nuevo: hace falta categoría (se crea si no existe) y precio de venta.
                categoria_nombre = _norm(val(fila, "Categoria"))
                categoria = None
                if categoria_nombre:
                    categoria = categorias_cache.get(categoria_nombre.lower())
                    if categoria is None:
                        categoria = models.Categoria(nombre=categoria_nombre)
                        db.add(categoria)
                        db.flush()
                        categorias_cache[categoria_nombre.lower()] = categoria

                precio_raw = val(fila, "PrecioVenta")
                if precio_raw in (None, ""):
                    errores.append({
                        "fila": n, "producto": nombre,
                        "motivo": "Producto nuevo sin PrecioVenta (obligatorio para altas).",
                    })
                    continue
                precio_venta = Decimal(str(precio_raw))

                producto = models.Producto(
                    nombre=nombre,
                    categoria_id=categoria.id if categoria else None,
                    precio_venta=precio_venta,
                    costo=costo_final,
                    mix_pct=0,
                    activo=True,
                )
                db.add(producto)
                db.flush()
                productos_cache[clave] = producto
                productos_nuevos_ids.append(producto.id)

                variante_id, error_variante = resolver_variante_para_fila(producto, atributos_fila)
                if error_variante:
                    # No debería pasar para un producto recién creado (el conjunto de atributos
                    # requerido se define a partir de esta misma fila), pero se cubre por las dudas.
                    errores.append({"fila": n, "producto": nombre, "motivo": error_variante})
                    continue

                db.add(models.Compra(
                    producto_id=producto.id, variante_id=variante_id, fecha=fecha, cantidad=cantidad,
                    costo_unitario=costo_final, proveedor="Importación Excel",
                ))
                db.flush()

                productos_creados.append({
                    "producto_id": producto.id, "producto": nombre,
                    "categoria": categoria.nombre if categoria else "Sin categoría",
                    "stock_inicial": cantidad, "costo": float(costo_final), "precio_venta": float(precio_venta),
                    "variante_id": variante_id,
                    "variante_descripcion": descripcion_variante(variante_id),
                })
            else:
                variante_id, error_variante = resolver_variante_para_fila(producto, atributos_fila)
                if error_variante:
                    errores.append({"fila": n, "producto": nombre, "motivo": error_variante})
                    continue

                # Reposición de un producto que ya existe en el catálogo.
                costo_promedio_antes = float(producto.costo)
                ultima_compra_previa = (
                    db.query(models.Compra)
                    .filter(models.Compra.producto_id == producto.id)
                    .order_by(models.Compra.fecha.desc(), models.Compra.id.desc())
                    .first()
                )
                costo_ultima_compra = float(ultima_compra_previa.costo_unitario) if ultima_compra_previa else None

                db.add(models.Compra(
                    producto_id=producto.id, variante_id=variante_id, fecha=fecha, cantidad=cantidad,
                    costo_unitario=costo_final, proveedor="Importación Excel",
                ))
                db.flush()
                calculations.recalcular_costo_promedio(db, producto.id)
                db.refresh(producto)
                costo_promedio_despues = float(producto.costo)

                compras_registradas.append({
                    "producto_id": producto.id, "producto": nombre, "cantidad": cantidad,
                    "costo_unitario": float(costo_final), "fecha": fecha.isoformat(),
                    "variante_id": variante_id,
                    "variante_descripcion": descripcion_variante(variante_id),
                })

                # El aviso se dispara comparando contra el costo de reposición (última compra
                # previa), no contra el promedio ponderado — ver simular_compra() en calculations.py.
                diferencia_vs_promedio_pct = (
                    round((costo_promedio_despues - costo_promedio_antes) / costo_promedio_antes * 100, 2)
                    if costo_promedio_antes else 0.0
                )
                diferencia_vs_ultima_compra_pct = (
                    round((float(costo_final) - costo_ultima_compra) / costo_ultima_compra * 100, 2)
                    if costo_ultima_compra else None
                )
                if diferencia_vs_ultima_compra_pct is not None and abs(diferencia_vs_ultima_compra_pct) > umbral_cambio_costo_pct:
                    precio_actual = float(producto.precio_venta)
                    cambios_costo.append({
                        "producto_id": producto.id,
                        "producto": nombre,
                        "costo_anterior": costo_promedio_antes,
                        "costo_nuevo": costo_promedio_despues,
                        "costo_ultima_compra": costo_ultima_compra,
                        "diferencia_vs_ultima_compra_pct": diferencia_vs_ultima_compra_pct,
                        "diferencia_vs_promedio_pct": diferencia_vs_promedio_pct,
                        "precio_venta_actual": precio_actual,
                        "precio_venta_sugerido": round(precio_actual * (1 + diferencia_vs_ultima_compra_pct / 100), 2),
                    })

        except (InvalidOperation, ValueError) as e:
            errores.append({"fila": n, "producto": nombre, "motivo": f"Dato inválido ({e})."})
            continue

    db.commit()

    # por si el mismo producto nuevo aparece en más de una fila de la planilla
    for pid in productos_nuevos_ids:
        calculations.recalcular_costo_promedio(db, pid)

    return {
        "productos_creados": productos_creados,
        "compras_registradas": compras_registradas,
        "cambios_costo": cambios_costo,
        "errores": errores,
    }
